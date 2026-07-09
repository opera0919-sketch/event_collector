# -*- coding: utf-8 -*-
"""오프라인 검증 — Gemini/증권사 사이트 실호출 없이 v2 정규화·동기화 로직을 검사.

실행: python tests/test_offline.py
(API 소진 0회 정책: vision 은 가짜 응답으로 대체, db 쓰기는 기록만 한다)
"""

import datetime as dt
import os
import sys

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))
os.environ.pop("GEMINI_API_KEY", None)  # 실호출 원천 차단

from pension_monitor import db, normalize, report, vision
from pension_monitor.classify import (
    content_hash, detect_accounts, source_event_id, source_content_hash,
    clip_detail, weekday_conflicts, infer_end_from_hold,
)


def test_clean_rows_junk_and_grounding():
    res = {
        "benefits": [
            {"condition": "IRP 순입금 1백만원 이상", "reward": "신세계 모바일상품권 2만원",
             "method": "전원", "limit_count": 0},
            {"condition": "참여 시", "reward": "No benefits information found",
             "method": "기타", "limit_count": 0},                      # EN 정크 → 제거
            {"condition": "가입 시", "reward": "혜택 내용 없음", "method": "기타",
             "limit_count": 0},                                        # KR 정크 → 제거
            {"condition": "이벤트 참여", "reward": "즐거운 투자 경험", "method": "전원",
             "limit_count": 0},                                        # 실질 토큰 없음 → 제거
        ],
        "conditions": [
            {"label": "대상", "value": "IRP 계좌 보유 고객"},
            {"label": "기타", "value": "명시되어 있지 않음"},           # 정크 → 제거
        ],
    }
    src = "이벤트 안내: IRP 순입금 1백만원 이상 시 신세계 모바일상품권 2만원 지급"
    b, c, grounded = normalize.clean_rows(res, source_text=src)
    assert len(b) == 1 and b[0]["benefit_text"] == "신세계 모바일상품권 2만원", b
    assert grounded, "원문에 있는 수치는 근거 일치여야 함"
    assert len(c) == 1 and c[0]["label"] == "대상", c
    # 근거 불일치: 원문에 없는 금액
    res2 = {"benefits": [{"condition": "순입금 5천만원 이상", "reward": "상품권 99만원",
                          "method": "전원", "limit_count": 0}], "conditions": []}
    b2, _, grounded2 = normalize.clean_rows(res2, source_text=src)
    assert len(b2) == 1 and not grounded2, "원문에 없는 수치는 근거 불일치"
    print("OK clean_rows (정크 한/영 차단, 실질 토큰, 근거 대조)")


def test_render():
    rows = [
        {"tier_no": 1, "condition_text": "순입금 1백만원 이상", "benefit_text": "상품권 1만원",
         "award_method": "전원", "award_limit": None, "source": "llm-text"},
        {"tier_no": 2, "condition_text": "디폴트옵션 지정", "benefit_text": "아메리카노 1잔",
         "award_method": "추첨", "award_limit": 3000, "source": "llm-text"},
    ]
    md = normalize.render_benefits(rows)
    assert md == "순입금 1백만원 이상 → 상품권 1만원 (전원)\n디폴트옵션 지정 → 아메리카노 1잔 (추첨 3,000명)", md
    cond = normalize.render_conditions([{"ord": 1, "label": "대상", "value_text": "IRP 보유", "source": "x"},
                                        {"ord": 2, "label": "기타", "value_text": "신청 필수", "source": "x"}])
    assert cond == "대상: IRP 보유\n신청 필수", cond
    print("OK render (캐노니컬 텍스트)")


def test_reconcile_period():
    year = dt.date.today().year
    # 1) 신뢰 목록(미래에셋) → list 유지, LLM 불일치는 검토 플래그만
    ev = {"firm_name": "미래에셋증권", "start_date": f"{year}-07-01", "end_date": f"{year}-09-30"}
    normalize.reconcile_period(ev, {"period_start": f"{year}-07-01", "period_end": f"{year}-10-31"})
    assert ev["date_source"] == "list" and ev["end_date"] == f"{year}-09-30"
    assert ev.get("needs_review") and "기간 불일치" in ev["review_reason"]
    # 2) KB(목록 불신) + 상세 본문 '기간' → detail
    ev = {"firm_name": "KB증권", "start_date": f"{year}-06-30", "end_date": f"{year}-07-01",
          "_detail_text": f"이벤트 기간 : {year}.04.01 ~ {year}.12.31 어쩌고"}
    normalize.reconcile_period(ev, {})
    assert (ev["start_date"], ev["end_date"], ev["date_source"]) == (f"{year}-04-01", f"{year}-12-31", "detail"), ev
    # 3) 상세도 없으면 LLM (유효 범위 내)
    ev = {"firm_name": "KB증권", "start_date": None, "end_date": None, "_detail_text": ""}
    normalize.reconcile_period(ev, {"period_start": f"{year}-05-01", "period_end": f"{year}-08-31"})
    assert ev["date_source"] == "llm" and ev["start_date"] == f"{year}-05-01"
    # 4) 어떤 출처로도 확신 없음 + 의심 목록 날짜 → 비우고 검토 플래그
    ev = {"firm_name": "KB증권", "start_date": "2024-10-25", "end_date": "2024-10-25", "_detail_text": ""}
    normalize.reconcile_period(ev, {})
    assert ev["start_date"] is None and ev["end_date"] is None
    assert ev.get("needs_review") and "기간 미확인" in ev["review_reason"]
    print("OK reconcile_period (신뢰원 규칙: list > detail > llm > null+검토)")


def test_accounts_conservative():
    acct = detect_accounts("연금 이벤트 대박")           # '연금' 단독 → 추정 금지
    assert not any([acct["acct_pension"], acct["acct_irp"], acct["acct_dc"]])
    acct = detect_accounts("퇴직연금 이벤트")            # 통칭 유지 (의미상 동치)
    assert acct["acct_irp"] and acct["acct_dc"]
    ev = {"firm_name": "X", "event_name": "e", "acct_pension": False, "acct_irp": False,
          "acct_dc": False, "acct_etc": None}
    normalize.apply_accounts(ev, {})
    assert ev.get("needs_review") and "대상계좌 미확인" in ev["review_reason"]
    print("OK accounts (보수 판정 + 미확인 플래그)")


def test_normalize_no_regression_and_cache():
    year = dt.date.today().year
    old = {"id": 7, "firm_name": "NH투자증권", "event_name": "IRP 이벤트",
           "start_date": f"{year}-05-01", "end_date": f"{year}-08-31",
           "source_event_id": "1000", "needs_review": False,
           "benefits": "순입금 1백만원 이상 → 상품권 1만원 (전원)",
           "conditions": "대상: IRP", "acct_irp": True, "extract_method": "text"}
    # 캐시 적중 (같은 sid + 같은 종료일 + 캐노니컬 양호) — vision 미설정이라 호출 자체가 없어야 함
    ev = {"firm_name": "NH투자증권", "event_name": "IRP 이벤트(개칭)",
          "source_event_id": "1000", "start_date": f"{year}-05-01", "end_date": f"{year}-08-31",
          "acct_pension": False, "acct_irp": False, "acct_dc": False, "acct_etc": None,
          "benefits": None, "conditions": None, "_detail_text": "본문" * 100}
    normalize.normalize_events([ev], [old])
    assert ev["benefits"] == old["benefits"] and ev["acct_irp"] is True
    assert not ev.get("rows_fresh")
    # 무회귀: 추출 불가(신규 아님) → 기존 값 유지 대신, 기존도 없으면 hint 폴백 + 검토
    ev2 = {"firm_name": "KB증권", "event_name": "새 이벤트", "source_event_id": "999",
           "start_date": None, "end_date": None, "acct_pension": True, "acct_irp": False,
           "acct_dc": False, "acct_etc": None, "benefits": None, "conditions": None,
           "_detail_text": "", "_benefits_hint": "순입금하면 상품권"}
    normalize.normalize_events([ev2], [])
    assert ev2["benefits"] == "순입금하면 상품권" and ev2["extract_method"] == "hint"
    assert ev2.get("needs_review") and "목록 요약 폴백" in ev2["review_reason"]
    print("OK normalize (캐시 재사용·무회귀·hint 폴백 검토 플래그)")


def test_normalize_rejects_junk_reused_from_old():
    """실전 재현(2026-07-06, 한투 id=43): 파이프라인 밖(구코드)에서 DB의 old.benefits
    가 '혜택 없음 (자료 없음)' 같은 정크로 재오염된 뒤, 오늘 실행의 신선 추출도
    실패하면 G3(무회귀)가 '비어있지 않다'는 이유만으로 그 정크를 영구 재사용했다.
    _is_trustworthy 재검사로 이 경로가 차단되고 정상적으로 미확인 처리돼야 한다."""
    year = dt.date.today().year
    old = {"id": 43, "firm_name": "한국투자증권", "event_name": "퇴직연금 DC 신규 입금 이벤트",
           "start_date": f"{year}-07-01", "end_date": f"{year}-09-30",
           "source_event_id": "6730", "needs_review": True,
           "review_reason": "정크 정화(v2): 혜택 무의미 응답 제거",
           "benefits": "퇴직연금 DC 신규 입금 → 혜택 없음 (자료 없음)",   # 구코드 재오염분
           "conditions": None, "extract_method": None}
    ev = {"firm_name": "한국투자증권", "event_name": "퇴직연금 DC 신규 입금 이벤트",
          "source_event_id": "6730", "start_date": f"{year}-07-01", "end_date": f"{year}-09-30",
          "acct_pension": False, "acct_irp": False, "acct_dc": True, "acct_etc": None,
          "benefits": None, "conditions": None, "_detail_text": ""}   # 오늘도 추출 실패
    normalize.normalize_events([ev], [old])
    assert ev["benefits"] is None, ev["benefits"]           # 정크가 재사용되면 안 됨
    assert ev["extract_method"] == "none"
    assert "혜택 미확인" in ev["review_reason"]
    print("OK normalize (구코드 재오염 정크의 무회귀 재사용 차단)")


def test_sync_v2():
    year = dt.date.today().year
    today = dt.date.today().isoformat()
    calls = {"patch": [], "post": [], "delete": []}
    db.fetch_all_events = lambda: existing
    db.enabled = lambda: True
    db._patch = lambda path, params, payload: calls["patch"].append((path, params, payload))
    db._post = lambda path, payload, prefer=None: (calls["post"].append((path, payload)), [{"id": 99}])[1]
    db._delete = lambda path, params: calls["delete"].append((path, params))

    existing = [
        {"id": 1, "firm_name": "NH투자증권", "event_name": "IRP 이벤트",
         "source_event_id": "1000", "start_date": f"{year}-05-01", "end_date": f"{year}-08-31",
         "status": "진행중", "missed_count": 0, "benefits": "옛 혜택 → 상품권 (전원)",
         "conditions": None, "content_hash": "old", "last_seen_at": f"{year}-06-30T00:00:00+00:00"},
        {"id": 2, "firm_name": "한국투자증권", "event_name": "만기 이벤트",
         "source_event_id": "2", "start_date": f"{year}-04-01", "end_date": "2020-06-30",
         "status": "진행중", "missed_count": 0, "content_hash": "x",
         "last_seen_at": f"{year}-06-29T00:00:00+00:00"},
    ]
    ev = {"firm_name": "NH투자증권", "event_name": "IRP 이벤트", "source_event_id": "1000",
          "start_date": f"{year}-05-01", "end_date": f"{year}-08-31", "status": None,
          "benefits": "새 혜택 → 상품권 2만원 (전원)", "conditions": "대상: IRP",
          "rows_fresh": True,
          "benefit_rows": [{"tier_no": 1, "condition_text": "새 혜택", "benefit_text": "상품권 2만원",
                            "award_method": "전원", "award_limit": None, "source": "llm-text"}],
          "condition_rows": [{"ord": 1, "label": "대상", "value_text": "IRP", "source": "llm-text"}],
          "needs_review": False, "review_reason": None, "extract_method": "text",
          "date_source": "list", "last_verified_at": "now"}
    ev["content_hash"] = content_hash(ev)
    diff = db.sync([ev], firms_failed=["한국투자증권"], trigger_type="manual")
    # 자식 테이블 교체 확인
    assert any(p[0] == "event_benefits" for p in calls["delete"]), calls["delete"]
    assert any(p[0] == "event_benefits" and p[1][0]["event_id"] == 1 for p in calls["post"])
    # 콘텐츠 변경은 rows_fresh 라서 '변경' 기록됨
    assert any(f == "benefits" for _, f, _, _ in diff["changed"])
    # 실패 증권사라도 만기(2020) 건은 종료
    assert [e["id"] for e in diff["closed"]] == [2]
    print("OK db.sync v2 (자식 교체, 신선 변경 기록, 만기 스윕):",
          {k: len(v) for k, v in diff.items() if isinstance(v, list)})

    md = report.build_report(diff, ["한국투자증권"])
    assert "새 혜택 → 상품권 2만원" in md
    assert "데이터 신뢰도" in md and "전건 정상" in md   # ev 는 needs_review=False
    assert "| ✅ |" in md                                # v2 표의 '확인' 컬럼
    print("OK report render (v2: 데이터 신뢰도 롤업 + 확인 컬럼)")

    # 회귀: 신규 INSERT 시 NOT NULL boolean 컬럼(needs_review 등)이 null 로 나가
    # 23502 로 조용히 실패하던 버그 (2026-07-03 실측) — 반드시 bool 로 강제돼야 함
    calls["post"].clear()
    new_ev = {"firm_name": "NH투자증권", "event_name": "신규 이벤트", "source_event_id": "9999",
              "start_date": f"{year}-07-01", "end_date": f"{year}-09-30",
              "benefits": "b → r", "conditions": None}   # needs_review/acct_* 미설정(None)
    new_ev["content_hash"] = content_hash(new_ev)
    db.sync([new_ev], firms_failed=[], trigger_type="manual")
    row = next(p[1] for p in calls["post"] if p[0] == "pension_events")
    for f in ("acct_pension", "acct_irp", "acct_dc", "needs_review"):
        assert row[f] is False, (f, row[f])
    print("OK insert bool 강제 (needs_review null 회귀 방지)")


def test_review_reason_cleared_on_success():
    """실전 재현(id=39, 2026-07-06): needs_review=False 인데 review_reason 이
    이전 값으로 남아있어, 새 xlsx '검토사유' 컬럼에 낡은 사유가 그대로 노출됐다.
    성공 추출(_apply_success)과 캐시 재사용(_reuse_cached, old 가 검토 불필요인
    경우) 모두 review_reason 을 None 으로 정리해야 한다."""
    # 1) 신선 추출 성공 경로 — vision 이 오프라인이라 정규화 파이프라인을 거치지
    #    않고 _apply_success 를 직접 검증 (test_typed_condition_columns 와 동일 방식)
    ev = {"firm_name": "KB증권", "event_name": "e", "review_reason": "낡은 사유(이전 실행)"}
    b = [{"tier_no": 1, "condition_text": "가입 시", "benefit_text": "수수료 평생 무료",
          "award_method": "전원", "award_limit": None, "source": "llm-text"}]
    normalize._apply_success(ev, b, [], grounded=True, method="ocr")
    assert ev["review_reason"] is None, ev["review_reason"]

    # 2) 캐시 재사용 경로 — old 가 검토 불필요(needs_review=False)인데 review_reason
    #    이 남아있는 레코드를 재사용할 때도 ev 에 그 사유를 물려주면 안 됨
    year = dt.date.today().year
    old = {"id": 39, "firm_name": "KB증권", "event_name": "IRP 무료 수수료 혜택 이벤트",
           "source_event_id": "10007688", "start_date": f"{year}-01-01", "end_date": f"{year}-12-31",
           "needs_review": False, "review_reason": "혜택 미확인(본문/이미지 추출 실패) — 원문 확인 필요",
           "benefits": "가입 시 → 수수료 평생 무료 (전원)", "conditions": None}
    ev2 = {"firm_name": "KB증권", "event_name": "IRP 무료 수수료 혜택 이벤트",
           "source_event_id": "10007688", "start_date": f"{year}-01-01", "end_date": f"{year}-12-31",
           "acct_pension": False, "acct_irp": True, "acct_dc": False, "acct_etc": None,
           "benefits": None, "conditions": None, "_detail_text": "본문" * 100}
    normalize.normalize_events([ev2], [old])
    assert ev2["benefits"] == old["benefits"]     # 캐시 재사용 확인
    assert ev2.get("review_reason") is None, ev2.get("review_reason")
    print("OK review_reason 정리 (성공 추출·캐시 재사용 모두 낡은 사유 잔존 방지)")


def test_report_review_rollup_and_xlsx_sheets():
    """v2 리포트/xlsx 가 검증 메타·조건 타입드 컬럼·자식 테이블을 실제로 노출하는지."""
    year = dt.date.today().year
    active_ok = {"firm_name": "NH투자증권", "event_name": "정상 이벤트", "id": 1,
                 "start_date": f"{year}-05-01", "end_date": f"{year}-08-31", "status": "진행중",
                 "acct_pension": True, "acct_irp": False, "acct_dc": False, "acct_etc": None,
                 "benefits": "조건 → 리워드 (전원)", "needs_review": False,
                 "extract_method": "text", "annual_cap_krw": 30000, "apply_required": True}
    active_review = {"firm_name": "KB증권", "event_name": "검토 이벤트", "id": 2,
                      "start_date": None, "end_date": None, "status": "진행중",
                      "acct_pension": False, "acct_irp": True, "acct_dc": False, "acct_etc": None,
                      "benefits": None, "remarks": "기간 미확인(목록 날짜 게시일 오인 의심)",
                      "needs_review": True, "review_reason": "기간 미확인(목록 날짜 게시일 오인 의심)",
                      "extract_method": "none"}
    diff = {"new": [], "closed": [], "changed": [], "active": [active_ok, active_review]}
    md = report.build_report(diff, [])
    assert "검토 필요 1건 (기간 미확인 1건)" in md, md
    assert "연간 혜택한도 명시 이벤트: 1건" in md
    assert "신청 필수 이벤트: 1건" in md
    assert "| ⚠ |" in md and "| ✅ |" in md
    assert "「검토 이벤트」 — 기간 미확인" in md

    xlsx = report.build_xlsx(
        [active_ok, active_review],
        benefit_rows=[{"event_id": 1, "tier_no": 1, "condition_text": "조건",
                       "benefit_text": "리워드", "award_method": "전원",
                       "award_limit": None, "source": "llm-text"}],
        condition_rows=[{"event_id": 1, "ord": 1, "label": "대상",
                         "value_text": "IRP 보유", "source": "llm-text"}])
    from openpyxl import load_workbook
    wb = load_workbook(io_bytes(xlsx))
    assert wb.sheetnames == ["이벤트 요약", "혜택 상세", "참여조건 상세"], wb.sheetnames
    assert wb["혜택 상세"].cell(row=2, column=1).value == "NH투자증권"   # event_id 조인 확인
    assert wb["참여조건 상세"].cell(row=2, column=4).value == "대상"
    print("OK report v2 (신뢰도 롤업·검토 사유 버킷·인사이트·xlsx 3시트+자식조인)")


def io_bytes(b):
    import io
    return io.BytesIO(b)


def test_typed_condition_columns():
    """_apply_success 가 조건 라벨 행 → 타입드 컬럼을 채우고 '기간' 행을 제외하는지."""
    ev = {"firm_name": "NH투자증권", "event_name": "e"}
    b = [{"tier_no": 1, "condition_text": "순입금 1백만원 이상", "benefit_text": "상품권 1만원",
          "award_method": "전원", "award_limit": None, "source": "llm-text"}]
    c = [{"ord": 1, "label": "대상", "value_text": "DC 최초 입금 가입자 (법인고객 제외)", "source": "llm-text"},
         {"ord": 2, "label": "기간", "value_text": "2026.07.01 ~ 2026.09.30", "source": "llm-text"},
         {"ord": 3, "label": "신청", "value_text": "별도 신청 없이 자동 참여", "source": "llm-text"},
         {"ord": 4, "label": "한도", "value_text": "퇴직연금 특별이익 제공한도(연간 누적 3만원) 내", "source": "llm-text"}]
    normalize._apply_success(ev, b, c, grounded=True, method="text")
    assert ev["eligibility"] == "DC 최초 입금 가입자 (법인고객 제외)"
    assert ev["exclusions"] == "법인고객 제외"          # §7-1: 제외 절 분리
    assert ev["apply_required"] is False
    assert ev["annual_cap_krw"] == 30000
    # '기간' 라벨은 conditions/자식 행에서 제외 (start/end 컬럼과 중복 금지)
    assert all(r["label"] != "기간" for r in ev["condition_rows"])
    assert "기간" not in ev["conditions"]
    assert [r["ord"] for r in ev["condition_rows"]] == [1, 2, 3]
    print("OK typed condition columns (파이프라인 연결 + 기간 중복 배제)")


def test_vision_image_parts():
    from pension_monitor import vision
    # 스크린샷(b64 dict) → inline part 직결 (네트워크 불필요)
    p = vision._image_part({"b64": "QUJD", "mime": "image/jpeg"})
    assert p == {"inline_data": {"mime_type": "image/jpeg", "data": "QUJD"}}, p
    assert vision._image_part({"b64": ""}) is None
    assert vision._image_part("not-a-url") is None
    print("OK vision._image_part (스크린샷 b64 / URL 분기)")


def test_source_event_id():
    assert source_event_id({"event_url": "https://x/go.able?linkcd=a&seq=10009676&idt=1"}) == "10009676"
    assert source_event_id({"event_url": "https://m.nhsec.com/customer/event/eventView?mNo=971"}) == "971"
    assert source_event_id({"event_url": "https://x/v01.do?cs_ecis_id=202603005&mod=S"}) == "202603005"
    assert source_event_id({"event_url": "https://x/noticeEvent.do?cmd=eventView&MenuSeqNo=3808"}) == "3808"
    assert source_event_id({"event_url": "https://x/Event.jsp?num=6711", "_detail_id": "6711"}) == "6711"
    assert source_event_id({"event_url": "https://x/mki7000/r01.do"}) is None
    print("OK source_event_id")


def test_clip_detail_preserves_tail():
    """P0-3: 상단 네비가 길어도 하단 유의사항(배수·제외재원) 꼬리는 절단 후에도 보존."""
    head = "메뉴\n" + ("본문내용가나다 " * 2000)          # 예산 초과할 만큼 긴 머리
    tail = "※ 유의사항\n타사이전 시 실적 1.5배 인정\n퇴직금은 제외됩니다"
    clipped = clip_detail(head + "\n" + tail, limit=8000)
    assert len(clipped) <= 8000
    assert "1.5배" in clipped and "제외됩니다" in clipped, "꼬리 소실"
    # 예산 내면 그대로
    assert clip_detail("짧은 본문") == "짧은 본문"
    print("OK clip_detail (유의사항 꼬리 보존)")


def test_weekday_conflicts_g6():
    """G6: 본문 'YYYY.MM.DD(요일)' 표기와 실제 요일 불일치 탐지 (E6 시즌 기간 오추출)."""
    # 2026-06-30 은 화요일 → (수) 표기는 불일치
    bad = weekday_conflicts("이벤트 종료일 2026. 06. 30(수) 까지")
    assert bad and "2026-06-30" in bad[0], bad
    assert weekday_conflicts("행사 2026.06.30(화) 마감") == []      # 올바른 요일은 통과
    assert weekday_conflicts("요일 표기 없음 2026.06.30") == []
    print("OK weekday_conflicts (G6 요일 정합성)")


def test_infer_end_from_hold_g9():
    """G9: 잔고유지 시작일 - 1일 = 이벤트 종료일 역산 (E5 KB 시즌3 기간 NULL 해결)."""
    assert infer_end_from_hold("잔고유지기간 2026.10.1 ~ 2026.10.31", 2026) == "2026-09-30"
    assert infer_end_from_hold("유지 기간 10월 1일부터", 2026) == "2026-09-30"
    assert infer_end_from_hold("혜택 안내만 있음", 2026) is None
    print("OK infer_end_from_hold (G9 잔고유지 역산)")


def test_clean_multipliers():
    """P0-1: 배수 배열 정규화 — scope(인정금액/리워드금액) 분리, 무효 배수 제거."""
    res = {"multipliers": [
        {"source_type": "타사이전", "multiplier": 1.5, "scope": "인정금액",
         "min_threshold_krw": 10000000, "extra_condition": ""},
        {"source_type": "비대면최초신규", "multiplier": 2, "scope": "리워드금액",
         "min_threshold_krw": 0, "extra_condition": "WRAP 가입"},
        {"source_type": "엉뚱", "multiplier": 0, "scope": "x",
         "min_threshold_krw": "bad", "extra_condition": ""},          # 무효 → 제거
    ]}
    rows = normalize.clean_multipliers(res, "llm-text")
    assert len(rows) == 2, rows
    assert rows[0]["scope"] == "인정금액" and rows[0]["multiplier"] == 1.5
    assert rows[0]["min_threshold_krw"] == 10000000
    assert rows[1]["scope"] == "리워드금액" and rows[1]["source_type"] == "비대면최초신규"
    print("OK clean_multipliers (배수 scope 분리·무효 제거)")


def test_source_content_hash_reextract():
    """P0-5: 상세 본문/이미지가 바뀌면 해시가 바뀌어 재추출을 유발 (누락 영구화 방지)."""
    a = source_content_hash({"_detail_text": "본문A", "_image_urls": ["u1"]})
    b = source_content_hash({"_detail_text": "본문B", "_image_urls": ["u1"]})
    c = source_content_hash({"_detail_text": "본문A", "_image_urls": ["u2"]})
    assert a != b and a != c
    assert a == source_content_hash({"_detail_text": "본문A", "_image_urls": ["u1"]})
    print("OK source_content_hash (원문 변경 → 재추출 트리거)")


def test_normalize_sets_reextract_meta():
    """P0-5: 정규화가 모든 이벤트에 source_content_hash + 스키마 버전을 부여."""
    ev = {"firm_name": "KB증권", "event_name": "e", "source_event_id": "1",
          "start_date": None, "end_date": None, "acct_pension": True, "acct_irp": False,
          "acct_dc": False, "acct_etc": None, "benefits": None, "conditions": None,
          "_detail_text": "상세 본문", "_image_urls": ["u1"]}
    normalize.normalize_events([ev], [])
    assert ev["extract_schema_version"] == vision.EXTRACT_SCHEMA_VERSION
    assert ev["source_content_hash"] == source_content_hash(
        {"_detail_text": "상세 본문", "_image_urls": ["u1"]})
    print("OK normalize 재추출 메타 부여")


def test_cache_invalidation_on_schema_bump():
    """P0-5: 종료일 동일 + benefits 양호 + needs_review=False 라도, 스키마 버전이
    낮으면 fast-cache 스킵이 아니라 재추출 경로를 타야 한다 (E3 누락 영구화 차단).
    vision 미설정이라 재추출은 실패하고 무회귀로 old 를 재사용하지만, 핵심은
    '캐시로 조용히 넘어가지 않고 재평가 대상이 된다'는 점이다."""
    year = dt.date.today().year
    old = {"id": 8, "firm_name": "삼성증권", "event_name": "연금 파워업",
           "source_event_id": "3808", "start_date": f"{year}-05-01", "end_date": f"{year}-07-31",
           "needs_review": False, "benefits": "이전 시 → 상품권 1만원 (전원)",
           "conditions": None, "source_content_hash": "STALE",
           "extract_schema_version": vision.EXTRACT_SCHEMA_VERSION - 1}
    ev = {"firm_name": "삼성증권", "event_name": "연금 파워업", "source_event_id": "3808",
          "start_date": f"{year}-05-01", "end_date": f"{year}-07-31",
          "acct_pension": True, "acct_irp": False, "acct_dc": False, "acct_etc": None,
          "benefits": None, "conditions": None, "_detail_text": ""}
    normalize.normalize_events([ev], [old])
    # 새 원문 해시/스키마 버전이 ev 에 반영됐고(재평가됨), 무회귀로 old benefits 유지
    assert ev["extract_schema_version"] == vision.EXTRACT_SCHEMA_VERSION
    assert ev["source_content_hash"] != old["source_content_hash"]
    assert ev["benefits"] == old["benefits"]
    assert not ev.get("rows_fresh")            # fast-cache 였다면 여기 도달 안 함
    print("OK 캐시 무효화 (스키마 버전 상향 시 재평가)")


def test_fast_cache_skips_extraction():
    """P0-5: 원문 해시 + 스키마 버전이 모두 일치하면 fast-cache 로 추출을 건너뛴다.
    (vision 을 활성화하고 호출 시 카운트 — 0 이어야 캐시 스킵 확인)."""
    year = dt.date.today().year
    det = "본문" * 100
    ev0 = {"_detail_text": det}
    old = {"id": 1, "firm_name": "NH투자증권", "event_name": "IRP", "source_event_id": "1",
           "start_date": f"{year}-05-01", "end_date": f"{year}-08-31", "needs_review": False,
           "benefits": "순입금 → 상품권 1만원 (전원)", "conditions": "대상: IRP", "acct_irp": True,
           "source_content_hash": source_content_hash(ev0),
           "extract_schema_version": vision.EXTRACT_SCHEMA_VERSION}
    called = {"n": 0}
    orig_enabled, orig_extract = vision.enabled, vision.extract_from_text
    vision.enabled = lambda: True
    vision.blocked = lambda: False
    vision.extract_from_text = lambda *a, **k: (called.__setitem__("n", called["n"] + 1), {})[1]
    try:
        ev = {"firm_name": "NH투자증권", "event_name": "IRP", "source_event_id": "1",
              "start_date": f"{year}-05-01", "end_date": f"{year}-08-31", "acct_pension": False,
              "acct_irp": False, "acct_dc": False, "acct_etc": None, "benefits": None,
              "conditions": None, "_detail_text": det}
        normalize.normalize_events([ev], [old])
    finally:
        vision.enabled, vision.extract_from_text = orig_enabled, orig_extract
    assert called["n"] == 0, "fast-cache 인데 추출을 호출함"
    assert ev["benefits"] == old["benefits"] and ev["acct_irp"] is True
    print("OK fast-cache 스킵 (원문/스키마 일치 시 추출 안 함)")


def test_check_coverage_g8():
    """G8: 이전 유치형인데 원문엔 배수·제외재원이 있으나 추출엔 없으면 검토 표기 (E1/E2)."""
    ev = {"firm_name": "삼성증권", "event_name": "타사 연금 이전 이벤트",
          "_detail_text": "타사에서 이전 시 실적 1.5배 인정. 퇴직금은 제외됩니다.",
          "benefits": "이전 시 → 상품권", "conditions": None}
    normalize.check_coverage(ev)
    assert ev.get("needs_review"), "배수/제외재원 누락이 잡혀야 함"
    assert "배수" in ev["review_reason"] and "제외재원" in ev["review_reason"]
    # 배수가 실제 추출되면 배수 플래그는 뜨지 않는다
    ev2 = {"firm_name": "삼성증권", "event_name": "타사 이전 이벤트",
           "_detail_text": "타사 이전 시 실적 1.5배 인정", "benefits": "이전 → 상품권",
           "multiplier_rows": [{"multiplier": 1.5}]}
    normalize.check_coverage(ev2)
    assert not ev2.get("needs_review")
    # 이전 유치형이 아니면 검사 스킵
    ev3 = {"firm_name": "NH투자증권", "event_name": "신규 가입 이벤트",
           "_detail_text": "신규 가입 시 상품권", "benefits": "x"}
    normalize.check_coverage(ev3)
    assert not ev3.get("needs_review")
    print("OK check_coverage (G8 필수 조항 커버리지)")


def test_check_period_collisions_g7():
    """G7: 같은 증권사에서 이름(시즌)은 다른데 기간이 동일하면 오추출 의심 표기 (E6)."""
    year = dt.date.today().year
    evs = [
        {"firm_name": "KB증권", "event_name": "TDF&ETF 시즌2",
         "start_date": f"{year}-04-01", "end_date": f"{year}-06-30"},
        {"firm_name": "KB증권", "event_name": "TDF&ETF 시즌3",
         "start_date": f"{year}-04-01", "end_date": f"{year}-06-30"},
    ]
    normalize.check_period_collisions(evs)
    assert all(e.get("needs_review") for e in evs)
    assert "기간 충돌" in evs[0]["review_reason"]
    # 기간이 다르면 충돌 없음
    ok = [{"firm_name": "KB증권", "event_name": "A 시즌1", "start_date": f"{year}-01-01",
           "end_date": f"{year}-02-01"},
          {"firm_name": "KB증권", "event_name": "A 시즌2", "start_date": f"{year}-03-01",
           "end_date": f"{year}-04-01"}]
    normalize.check_period_collisions(ok)
    assert not any(e.get("needs_review") for e in ok)
    print("OK check_period_collisions (G7 동일기간 충돌)")


def test_apply_success_multipliers_and_meta():
    """P0-1/P3: _apply_success 가 배수행·stackable·연간횟수 메타를 이벤트에 부착."""
    ev = {"firm_name": "KB증권", "event_name": "타사 이전 시즌3"}
    b = [{"tier_no": 1, "condition_text": "이전", "benefit_text": "상품권 2만원",
          "award_method": "전원", "award_limit": None, "source": "llm-text"}]
    mult = [{"source_type": "타사이전", "multiplier": 2, "scope": "인정금액",
             "min_threshold_krw": 0, "extra_condition": "", "source": "llm-text"},
            {"source_type": "비대면최초신규", "multiplier": 2, "scope": "리워드금액",
             "min_threshold_krw": 0, "extra_condition": "", "source": "llm-text"}]
    normalize._apply_success(ev, b, [], grounded=True, method="text",
                             mult_rows=mult, stackable=True, annual_claim_limit=1)
    assert ev["multiplier_rows"] == mult and len(ev["multiplier_rows"]) == 2
    assert ev["stackable"] is True and ev["annual_claim_limit"] == 1
    assert ev["rows_fresh"]
    print("OK _apply_success (배수·중복가능·연간횟수 메타)")


def test_marketing_consent_label_free_fallback_g10():
    """G10: 라벨/키워드로 못 잡아도 '개인(신용)정보 선택동의 필수' 원문 스캔으로 보충 (E7)."""
    from src.backfill_conditions import parse_conditions
    p = parse_conditions("본 이벤트 참여 시 개인(신용)정보 선택 동의 필수 입니다")
    assert p["marketing_consent_required"] is True, p
    # 부정 표현이 있으면 보충하지 않는다 (오탐 방지)
    p2 = parse_conditions("마케팅 동의 없이 자동 참여 가능")
    assert p2["marketing_consent_required"] is False
    print("OK marketing_consent 라벨무관 폴백 (G10)")


def test_xlsx_multiplier_sheet():
    """P3: multiplier_rows 를 넘기면 '배수 상세' 4번째 시트가 생성되고 event_id 로 조인."""
    ok = {"firm_name": "KB증권", "event_name": "타사 이전 시즌3", "id": 5, "status": "진행중",
          "acct_pension": True, "acct_irp": False, "acct_dc": False, "acct_etc": None,
          "benefits": "이전 → 상품권", "needs_review": False, "stackable": True,
          "annual_claim_limit": 1}
    xlsx = report.build_xlsx(
        [ok], benefit_rows=[], condition_rows=[],
        multiplier_rows=[{"event_id": 5, "source_type": "타사이전", "multiplier": 2,
                          "scope": "인정금액", "min_threshold_krw": 10000000,
                          "extra_condition": "", "source": "llm-text"}])
    from openpyxl import load_workbook
    wb = load_workbook(io_bytes(xlsx))
    assert wb.sheetnames == ["이벤트 요약", "혜택 상세", "참여조건 상세", "배수 상세"], wb.sheetnames
    ws = wb["배수 상세"]
    assert ws.cell(row=2, column=1).value == "KB증권"       # event_id=5 조인
    assert ws.cell(row=2, column=4).value == 2               # 배수 값
    # 하위 호환: multiplier_rows 미전달 시 3시트 유지
    wb3 = load_workbook(io_bytes(report.build_xlsx([ok], [], [])))
    assert "배수 상세" not in wb3.sheetnames
    print("OK xlsx 배수 상세 시트 (P3 자식 조인·하위호환)")


def test_excluded_firm_not_auto_closed():
    """P2-2: 수집 제외 증권사(키움)의 미만기 잔존 건은 '미노출 → 종료'가 아니라
    needs_review 표기 (E8). 만기 건은 종료하되 close_reason 을 남긴다."""
    year = dt.date.today().year
    calls = {"patch": []}
    db.fetch_all_events = lambda: existing
    db.enabled = lambda: True
    db._patch = lambda path, params, payload: calls["patch"].append((params, payload))
    db._post = lambda path, payload, prefer=None: [{"id": 1}]
    db._delete = lambda path, params: None
    existing = [
        {"id": 30, "firm_name": "키움증권", "event_name": "연금 이벤트(웹검색 잔존)",
         "start_date": f"{year}-01-01", "end_date": f"{year}-12-31", "status": "진행중",
         "missed_count": 1, "content_hash": "x", "last_seen_at": f"{year}-06-01T00:00:00+00:00"},
    ]
    diff = db.sync([], firms_failed=[], trigger_type="manual")
    # 미만기 키움 건은 종료되지 않음
    assert diff["closed"] == [], diff["closed"]
    # 대신 needs_review + 사유가 기록됨
    marked = [p for _, p in calls["patch"] if p.get("needs_review")]
    assert marked and "수집 제외" in marked[0]["review_reason"], calls["patch"]
    print("OK 수집 제외 증권사 격리 (E8 자동종료 차단)")


if __name__ == "__main__":
    test_clip_detail_preserves_tail()
    test_weekday_conflicts_g6()
    test_infer_end_from_hold_g9()
    test_clean_multipliers()
    test_source_content_hash_reextract()
    test_normalize_sets_reextract_meta()
    test_cache_invalidation_on_schema_bump()
    test_fast_cache_skips_extraction()
    test_check_coverage_g8()
    test_check_period_collisions_g7()
    test_apply_success_multipliers_and_meta()
    test_marketing_consent_label_free_fallback_g10()
    test_xlsx_multiplier_sheet()
    test_excluded_firm_not_auto_closed()
    test_clean_rows_junk_and_grounding()
    test_render()
    test_reconcile_period()
    test_accounts_conservative()
    test_normalize_no_regression_and_cache()
    test_normalize_rejects_junk_reused_from_old()
    test_review_reason_cleared_on_success()
    test_sync_v2()
    test_report_review_rollup_and_xlsx_sheets()
    test_typed_condition_columns()
    test_vision_image_parts()
    test_source_event_id()
    print("\n전체 오프라인 검증 통과 (외부 API 호출 0회)")
