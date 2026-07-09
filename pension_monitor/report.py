# -*- coding: utf-8 -*-
"""주간 리포트 생성 (v2 — 마스터/조건·혜택 이원화 스키마 반영).

메일 본문(마크다운)과 xlsx 첨부 모두, v1(단일 conditions/benefits 텍스트 블롭)
대신 검증 메타(needs_review/extract_method/date_source)와 조건 타입드 컬럼
(eligibility/annual_cap_krw 등), 혜택 티어·조건 자식 테이블을 실제로 노출한다.
"""

import datetime as dt
import io

from .config import FIRMS

_EXTRACT_LABELS = {"text": "본문 추출", "ocr": "이미지 OCR", "heuristic": "휴리스틱",
                   "hint": "목록 요약", "none": "미확인"}
_DATE_SRC_LABELS = {"list": "목록", "detail": "상세본문", "llm": "AI 추출"}


def _tri_label(v, yes, no):
    if v is True:
        return yes
    if v is False:
        return no
    return ""


def _cap_label(v):
    return f"{v:,}원" if isinstance(v, int) else ""


def _prep_row(r):
    """DB 원시 값(코드/불리언/None)을 xlsx 표시용 라벨로 변환한 사본."""
    out = dict(r)
    out["date_source_label"] = _DATE_SRC_LABELS.get(r.get("date_source"), r.get("date_source") or "")
    out["extract_method_label"] = _EXTRACT_LABELS.get(r.get("extract_method"), r.get("extract_method") or "")
    out["apply_required_label"] = _tri_label(r.get("apply_required"), "신청필수", "자동적용")
    out["marketing_consent_required_label"] = _tri_label(r.get("marketing_consent_required"), "필요", "불필요")
    out["annual_cap_krw_label"] = _cap_label(r.get("annual_cap_krw"))
    out["stackable_label"] = _tri_label(r.get("stackable"), "중복가능", "중복불가")
    acl = r.get("annual_claim_limit")
    out["annual_claim_limit_label"] = f"{acl}회" if isinstance(acl, int) and acl > 0 else ""
    out["needs_review_label"] = "⚠ 검토" if r.get("needs_review") else ""
    out["last_verified_at"] = (r.get("last_verified_at") or "")[:10]
    return out


# 메일 첨부 xlsx 시트 1: 이벤트 요약 (마스터 + 검증 메타 + 조건 타입드 컬럼)
_XLSX_COLS = [
    ("firm_name", "증권사"), ("event_name", "이벤트명"), ("status", "상태"),
    ("start_date", "시작일"), ("end_date", "종료일"), ("date_source_label", "기간출처"),
    ("acct_pension", "연금저축"), ("acct_irp", "IRP"), ("acct_dc", "DC"), ("acct_etc", "기타계좌"),
    ("eligibility", "대상"), ("exclusions", "제외대상"),
    ("apply_required_label", "신청"), ("marketing_consent_required_label", "마케팅동의"),
    ("annual_cap_krw_label", "연간한도"), ("hold_condition", "유지조건"),
    ("stackable_label", "중복수령"), ("annual_claim_limit_label", "연간횟수"),
    ("benefits", "혜택내용"), ("cond_notes", "기타조건"),
    ("needs_review_label", "검토필요"), ("review_reason", "검토사유"),
    ("extract_method_label", "추출방식"), ("last_verified_at", "최종검증"),
    ("event_url", "출처URL"),
]

# 시트 2: 혜택 상세(티어) — event_benefits 자식 테이블
_BENEFIT_XLSX_COLS = [
    ("firm_name", "증권사"), ("event_name", "이벤트명"), ("tier_no", "순번"),
    ("condition_text", "조건"), ("benefit_text", "혜택"),
    ("award_method", "지급방식"), ("award_limit", "인원한도"), ("source", "출처"),
]

# 시트 3: 참여조건 상세(라벨:값) — event_conditions 자식 테이블
_CONDITION_XLSX_COLS = [
    ("firm_name", "증권사"), ("event_name", "이벤트명"), ("ord", "순번"),
    ("label", "구분"), ("value_text", "내용"), ("source", "출처"),
]

# 시트 4: 배수(승수) 상세 — pension_event_multipliers 자식 테이블
_MULTIPLIER_XLSX_COLS = [
    ("firm_name", "증권사"), ("event_name", "이벤트명"), ("source_type", "재원/자격"),
    ("multiplier", "배수"), ("scope", "적용대상"), ("min_threshold_krw", "최소금액"),
    ("extra_condition", "추가요건"), ("source", "출처"),
]

_WIDTHS = {"이벤트명": 34, "대상": 44, "제외대상": 34, "유지조건": 34, "혜택내용": 60,
           "기타조건": 34, "검토사유": 34, "출처URL": 46, "조건": 32, "혜택": 28, "내용": 44}


def _write_sheet(ws, cols, rows):
    from openpyxl.styles import Font, Alignment
    ws.append([label for _, label in cols])
    for c in ws[1]:
        c.font = Font(bold=True)
    for r in rows:
        out = []
        for key, _ in cols:
            v = r.get(key)
            if isinstance(v, bool):
                v = "○" if v else ""
            out.append("" if v is None else v)
        ws.append(out)
    for i, (_, label) in enumerate(cols, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = _WIDTHS.get(label, 12)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="top", wrap_text=True)
    ws.freeze_panes = "A2"


def build_xlsx(rows: list, benefit_rows: list = None, condition_rows: list = None,
               multiplier_rows: list = None) -> bytes:
    """DB 테이블을 xlsx 시트로 직렬화: 이벤트 요약(마스터) + 혜택 상세(티어)
    + 참여조건 상세 (+ multiplier_rows 가 주어지면 배수 상세). benefit_rows/
    condition_rows/multiplier_rows 는 event_id 로 마스터와 조인하거나(DB 조회분),
    firm_name/event_name 을 이미 담고 있으면(로컬 폴백분) 그대로 사용한다."""
    from openpyxl import Workbook
    wb = Workbook()

    rows = sorted(rows or [], key=lambda e: (e.get("status") != "진행중",
                                             e.get("firm_name") or "", e.get("end_date") or "9999"))
    prepped = [_prep_row(r) for r in rows]
    ws = wb.active
    ws.title = "이벤트 요약"
    _write_sheet(ws, _XLSX_COLS, prepped)

    id_map = {r.get("id"): (r.get("firm_name"), r.get("event_name")) for r in rows if r.get("id")}

    def _join(children, cols):
        out = []
        for c in children or []:
            firm, name = c.get("firm_name"), c.get("event_name")
            if firm is None and name is None:
                firm, name = id_map.get(c.get("event_id"), (None, None))
            if firm is None and name is None:
                continue           # 마스터에서 이미 정리된(삭제된) 이벤트의 고아 행 — 표시 제외
            out.append({**c, "firm_name": firm, "event_name": name})
        sort_key = cols[2][0]      # tier_no 또는 ord
        out.sort(key=lambda r: (r.get("firm_name") or "", r.get("event_name") or "", r.get(sort_key) or 0))
        return out

    ws2 = wb.create_sheet("혜택 상세")
    _write_sheet(ws2, _BENEFIT_XLSX_COLS, _join(benefit_rows, _BENEFIT_XLSX_COLS))

    ws3 = wb.create_sheet("참여조건 상세")
    _write_sheet(ws3, _CONDITION_XLSX_COLS, _join(condition_rows, _CONDITION_XLSX_COLS))

    # 배수 상세는 데이터가 전달된 경우에만 시트를 추가 (하위 호환: 기존 3시트 유지)
    if multiplier_rows is not None:
        ws4 = wb.create_sheet("배수 상세")
        _write_sheet(ws4, _MULTIPLIER_XLSX_COLS, _join(multiplier_rows, _MULTIPLIER_XLSX_COLS))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _b(v):
    return "○" if v else ""


def _fmt_period(ev):
    s = ev.get("start_date") or "?"
    e = ev.get("end_date") or "상시"
    return f"{s} ~ {e}"


# review_reason 문구를 대분류로 묶어 '데이터 신뢰도' 롤업에 쓴다.
_REVIEW_BUCKETS = [
    ("근거 불일치", "근거 불일치"),
    ("기간 미확인", "기간 미확인"),
    ("기간 불일치", "기간 미확인"),
    ("혜택 미확인", "혜택 미확인"),
    ("목록 요약 폴백", "목록 요약 폴백"),
    ("대상계좌 미확인", "대상계좌 미확인"),
    ("재검증 실패", "재검증 실패"),
    ("정크 정화", "정크 정화(자동)"),
]


def _bucket(reason):
    for needle, label in _REVIEW_BUCKETS:
        if needle in (reason or ""):
            return label
    return "기타"


def build_report(diff: dict, firms_failed: list) -> str:
    today = dt.date.today()
    active = sorted(diff["active"], key=lambda e: (e["firm_name"], e.get("end_date") or "9999"))
    lines = [
        f"# 연금 이벤트 리포트 ({today.isoformat()} 기준)",
        "",
        "## 요약",
        f"진행중 {len(active)}건 | 🆕 신규 {len(diff['new'])}건 | "
        f"🔚 종료 {len(diff['closed'])}건 | ✏️ 변경 {len(diff['changed'])}건 (직전 대비)",
        "",
    ]
    if firms_failed:
        lines += [f"⚠️ 수집 실패: {', '.join(firms_failed)} — 직전 수집분을 아래 표에 유지 "
                  f"(⚠ 최종확인일 표기)", ""]

    # 데이터 신뢰도 롤업 — 검증 게이트(needs_review) 결과를 사유별로 집계해
    # 표를 한 줄씩 읽지 않아도 이번 리포트의 신뢰 수준을 바로 가늠하게 한다.
    review_active = [e for e in active if e.get("needs_review")]
    if review_active:
        buckets = {}
        for e in review_active:
            b = _bucket(e.get("review_reason"))
            buckets[b] = buckets.get(b, 0) + 1
        detail = ", ".join(f"{k} {v}건" for k, v in sorted(buckets.items(), key=lambda kv: -kv[1]))
        lines.append(f"📋 데이터 신뢰도: 정상 {len(active) - len(review_active)}건 / "
                     f"검토 필요 {len(review_active)}건 ({detail})")
    else:
        lines.append(f"📋 데이터 신뢰도: 전건 정상 ({len(active)}건)")
    lines.append("")

    lines.append("## 직전 대비 주요 변동")
    if not (diff["new"] or diff["closed"] or diff["changed"]):
        lines.append("- 변동 없음")
    for ev in diff["new"]:
        lines.append(f"- 🆕 {ev['firm_name']} 「{ev['event_name']}」 ({_fmt_period(ev)})")
    for ev in diff["closed"]:
        lines.append(f"- 🔚 {ev['firm_name']} 「{ev['event_name']}」 종료")
    for ev, f, o, n in diff["changed"]:
        o1 = str(o or "").replace("\n", " · ")[:80]
        n1 = str(n or "").replace("\n", " · ")[:80]
        lines.append(f"- ✏️ {ev['firm_name']} 「{ev['event_name']}」 {f}: {o1} → {n1}")
    lines.append("")

    soon = [e for e in active if e.get("end_date")
            and 0 <= (dt.date.fromisoformat(e["end_date"]) - today).days <= 7]
    lines.append("## 종료 임박 (7일 이내 마감)")
    if soon:
        for ev in sorted(soon, key=lambda e: e["end_date"]):
            dday = (dt.date.fromisoformat(ev["end_date"]) - today).days
            lines.append(f"- {ev['firm_name']} 「{ev['event_name']}」 {ev['end_date']} 마감 (D-{dday})")
    else:
        lines.append("- 해당 없음")
    lines.append("")

    lines += [
        "## 진행중 이벤트 현황 (증권사별)",
        "| 증권사 | 이벤트명 | 기간 | 연금저축 | IRP | DC | 기타 | 혜택 요약 | 확인 |",
        "|---|---|---|---|---|---|---|---|---|",
    ]
    for ev in active:
        # 표준 형식 혜택은 줄바꿈 포함 → 표 셀에선 ' · '로 접어 한 줄로(원문/xlsx는 줄바꿈 유지)
        benefits = (ev.get("benefits") or ev.get("remarks") or "")
        benefits = benefits.replace("\n", " · ").replace("|", "/")[:70]
        name = ev["event_name"][:40].replace("|", "/")
        url = ev.get("event_url") or ""
        name_cell = f"[{name}]({url})" if url.startswith("http") else name
        if ev.get("stale_seen"):
            name_cell += f" ⚠(최종확인 {ev['stale_seen']})"
        confirm = "⚠" if ev.get("needs_review") else "✅"
        lines.append(
            f"| {ev['firm_name']} | {name_cell} | {_fmt_period(ev)} "
            f"| {_b(ev.get('acct_pension'))} | {_b(ev.get('acct_irp'))} "
            f"| {_b(ev.get('acct_dc'))} | {ev.get('acct_etc') or ''} | {benefits} | {confirm} |")
    lines.append("")

    lines.append("## 인사이트")
    by_firm = {}
    for ev in active:
        by_firm.setdefault(ev["firm_name"], []).append(ev)
    irp_n = sum(1 for e in active if e.get("acct_irp"))
    ps_n = sum(1 for e in active if e.get("acct_pension"))
    most = max(by_firm.items(), key=lambda kv: len(kv[1]))[0] if by_firm else None
    none_firms = [f for f in FIRMS if f not in by_firm and f not in firms_failed]
    if most:
        lines.append(f"- 진행중 이벤트 최다 증권사: {most} ({len(by_firm[most])}건)")
    lines.append(f"- 대상계좌 분포: 연금저축 {ps_n}건 / IRP {irp_n}건")
    cap_n = sum(1 for e in active if e.get("annual_cap_krw"))
    if cap_n:
        lines.append(f"- 연간 혜택한도 명시 이벤트: {cap_n}건 (대부분 퇴직연금 감독규정 3만원)")
    apply_n = sum(1 for e in active if e.get("apply_required") is True)
    if apply_n:
        lines.append(f"- 신청 필수 이벤트: {apply_n}건 (자동 참여 아님 — 별도 신청 필요)")
    if diff["new"]:
        lines.append(f"- 신규 {len(diff['new'])}건 — "
                     + ", ".join(sorted({e['firm_name'] for e in diff['new']})))
    if none_firms:
        lines.append(f"- 연금 이벤트 미진행: {', '.join(none_firms)}")
    lines.append("")

    review = [e for e in active if e.get("remarks")]
    if review:
        lines.append("## 검토 필요")
        for ev in review:
            src = _EXTRACT_LABELS.get(ev.get("extract_method"), "")
            src_note = f" [{src}]" if src and src != "미확인" else ""
            lines.append(f"- {ev['firm_name']} 「{ev['event_name']}」 — {ev['remarks']}{src_note}")
        lines.append("")

    lines.append(f"---\n*자동 생성: pension_monitor / 데이터 출처: 각 증권사 공식 홈페이지*")
    return "\n".join(lines)
